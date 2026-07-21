"""
ETL: Educacion Provincia.xlsx → indicadores (anuario_educacion)

Lee el Excel con datos educativos de Córdoba por departamento (2024)
y genera INSERTs SQL para la tabla indicadores.

Estructura de salida por fila:
  - categoria: 'anuario_educacion'
  - indicador_nombre: ej. "Unidades educativas - Primario"
  - valor: numérico
  - unidad: 'casos' | '%'
  - periodo: 2024
  - region: nombre del departamento
  - desglose: JSONB con breakdowns (sector, genero, etc.)
  - fuente: 'Ministerio de Educación - Anuario Estadístico'
"""

import json
from pathlib import Path

EXCEL_PATH = r"C:\Users\20409409009\Desktop\01. Monitoreo\TABLEROS Y BASES\Datos y Bases\bbdd\Educacion Provincia.xlsx"
PERIODO = 2024
FUENTE = "Ministerio de Educación - Anuario Estadístico"
CATEGORIA = "anuario_educacion"

# ─── Mapeo de tabs a indicadores ─────────────────────────────

TAB_CONFIGS = [
    # (sheet_name, nivel, [
    #   (col_idx, indicador_nombre, unidad, desglose_fixed),
    #   ...
    # ])
    
    # ── Primario ──────────────────────────────────────────
    ("Primario", "Primario", [
        (1, "Unidades educativas", "casos", {"sector": "Total"}),
        (2, "Alumnos", "casos", {"sector": "Total"}),
        (3, "Personal docente", "casos", {"sector": "Total"}),
        (4, "Unidades educativas", "casos", {"sector": "Estatal"}),
        (5, "Alumnos", "casos", {"sector": "Estatal"}),
        (6, "Personal docente", "casos", {"sector": "Estatal"}),
        (7, "Unidades educativas", "casos", {"sector": "Privado"}),
        (8, "Alumnos", "casos", {"sector": "Privado"}),
        (9, "Personal docente", "casos", {"sector": "Privado"}),
    ]),
    
    # ── Repitentes y Sobreedad - Primario ─────────────────
    # NOTA: columna 0 (Alumnos total) se saltea porque ya existe en el tab principal
    ("Repitentes-sobreedad Primario", "Primario", [
        (2, "Repitentes", "casos", {}),
        (3, "Tasa de repitencia", "%", {}),
        (4, "Sobreedad", "casos", {}),
        (5, "Tasa de sobreedad", "%", {}),
    ]),
    
    # ── Docentes - Primario ──────────────────────────────
    ("Docentes Primario", "Primario", [
        (1, "Docentes", "casos", {"genero": "Mujeres"}),
        (2, "Docentes", "casos", {"genero": "Varones"}),
    ]),
    
    # ── Secundario ────────────────────────────────────────
    ("Secundario", "Secundario", [
        (1, "Unidades educativas", "casos", {"sector": "Total"}),
        (2, "Alumnos", "casos", {"sector": "Total"}),
        (3, "Personal docente", "casos", {"sector": "Total"}),
        (4, "Unidades educativas", "casos", {"sector": "Estatal"}),
        (5, "Alumnos", "casos", {"sector": "Estatal"}),
        (6, "Personal docente", "casos", {"sector": "Estatal"}),
        (7, "Unidades educativas", "casos", {"sector": "Privado"}),
        (8, "Alumnos", "casos", {"sector": "Privado"}),
        (9, "Personal docente", "casos", {"sector": "Privado"}),
    ]),
    
    # ── Repitentes y Sobreedad - Secundario ──────────────
    # NOTA: columna 0 (Alumnos total) se saltea porque ya existe en el tab principal
    ("Repitentes-sobred Secundario", "Secundario", [
        (2, "Repitentes", "casos", {}),
        (3, "Tasa de repitencia", "%", {}),
        (4, "Sobreedad", "casos", {}),
        (5, "Tasa de sobreedad", "%", {}),
    ]),
    
    # ── Docentes - Secundario ────────────────────────────
    ("Docentes Secundario", "Secundario", [
        (1, "Docentes", "casos", {"genero": "Mujeres"}),
        (2, "Docentes", "casos", {"genero": "Varones"}),
    ]),
]


def normalizar_departamento(nombre: str) -> str:
    """Limpia y normaliza nombres de departamento."""
    nombre = nombre.strip()
    # Correcciones de encoding
    replacements = {
        "C¢rdoba": "Córdoba",
        "C¾rdoba": "Córdoba",
        "Córdoba": "Córdoba",
        "Col¢n": "Colón",
        "Cruz del Eje": "Cruz del Eje",
        "General Roca": "General Roca",
        "General San Mart¡n": "General San Martín",
        "General San Mart?n": "General San Martín",
        "Ischil¡n": "Ischilín",
        "Ju¡rez Celman": "Juárez Celman",
        "Marcos Ju¡rez": "Marcos Juárez",
        "Pte. Roque S en Pe a": "Presidente Roque Sáenz Peña",
        "Pte. Roque Sáenz Peña": "Presidente Roque Sáenz Peña",
        "Pte. Roque Senz Pea": "Presidente Roque Sáenz Peña",
        "Punilla": "Punilla",
        "R¡o Cuarto": "Río Cuarto",
        "R¡o Primero": "Río Primero",
        "R¡o Seco": "Río Seco",
        "R¡o Segundo": "Río Segundo",
        "San Mart¡n": "San Martín",
    }
    for old, new in replacements.items():
        if old in nombre:
            nombre = nombre.replace(old, new)
    # Normalize common diacritics
    nombre = nombre.replace("¢", "ó").replace("¾", "ó").replace("¡", "í")
    return nombre.strip()


def valor_seguro(val):
    """Convierte un valor a número, tratando None, string vacío, etc."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        val = val.strip()
        if val == "" or val == "-":
            return 0
        try:
            return float(val.replace(",", "."))
        except ValueError:
            return 0
    return 0


def generate_inserts():
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    all_values = []
    seen = set()  # para evitar duplicados exactos
    
    for sheet_name, nivel, columns in TAB_CONFIGS:
        if sheet_name not in wb.sheetnames:
            print(f"  Sheet '{sheet_name}' not found, skipping")
            continue
        
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
        
        for row in rows:
            departamento_val = row[0]
            if departamento_val is None:
                continue
            region = normalizar_departamento(str(departamento_val))
            
            for col_idx, base_name, unidad, extra_desglose in columns:
                if col_idx >= len(row):
                    continue
                raw_val = row[col_idx]
                v = valor_seguro(raw_val)
                
                # Construir nombre de indicador
                if unidad == "%":
                    indicador_nombre = f"{base_name} - {nivel}"
                else:
                    indicador_nombre = f"{base_name} - {nivel}"
                
                # Desglose
                desglose = dict(extra_desglose)
                
                # Clave única para dedup
                key = (indicador_nombre, nivel, region, str(desglose))
                if key in seen:
                    continue
                seen.add(key)
                
                # Escape para SQL
                nombre_escaped = indicador_nombre.replace("'", "''")
                region_escaped = region.replace("'", "''")
                desglose_json = json.dumps(desglose, ensure_ascii=False)
                
                all_values.append(
                    f"('{nombre_escaped}', '{CATEGORIA}', {v}, '{unidad}', {PERIODO}, "
                    f"'{region_escaped}', '{desglose_json}'::jsonb, '{FUENTE}')"
                )
    
    return all_values


def main():
    import openpyxl
    print("Leyendo Educacion Provincia.xlsx...")
    
    values = generate_inserts()
    
    print(f"Generados {len(values)} registros")
    
    # Verificar: muestrear algunos
    print("\nSample de indicadores generados:")
    sample_nombres = sorted(set(v.split(", ")[0] for v in values[:50]))
    for n in sample_nombres[:10]:
        count = sum(1 for v in values if v.startswith(n))
        print(f"  {repr(n)} -> {count} filas")
    
    # Generar SQL
    output_path = Path(__file__).parent / "educacion_provincia_2024_load.sql"
    
    # Batch en inserts de 100 filas para no exceder límites
    batch_size = 100
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("-- Carga de datos Educacion Provincia 2024\n")
        f.write(f"-- Generado: {len(values)} registros\n")
        f.write(f"-- Categoria: {CATEGORIA}\n\n")
        
        for i in range(0, len(values), batch_size):
            batch = values[i:i + batch_size]
            f.write("INSERT INTO indicadores (indicador_nombre, categoria, valor, unidad, periodo, region, desglose, fuente) VALUES\n")
            f.write(",\n".join(batch))
            f.write(";\n\n")
    
    print(f"\nSQL escrito en: {output_path}")
    print(f"   Total INSERTS: {len(values)}")
    print(f"   Batches: {(len(values) + batch_size - 1) // batch_size}")


if __name__ == "__main__":
    main()
