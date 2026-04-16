#!/usr/bin/env python3
"""
ETL script para cargar datos de Inversión desde Excel a Supabase.

Reads: C:/Users/Usuario/Desktop/08. Monitoreo-20260416T133722Z-3-001/01. Monitoreo/TABLEROS Y BASES/Inversion/BASE DE DATOS.xlsx
Inserts into: indicadores table with categoria='inversion'
"""

import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

# Configuration
EXCEL_FILE = Path("C:/Users/Usuario/Desktop/08. Monitoreo-20260416T133722Z-3-001/01. Monitoreo/TABLEROS Y BASES/Inversion/BASE DE DATOS.xlsx")
SHEET_NAME = "BASE DE DATOS"

# Categories to extract (childhood-related investments)
CATEGORIES = {
    "Educaci\u00f3n": "Inversi\u00f3n Educaci\u00f3n",
    "Salud": "Inversi\u00f3n Salud",
    "Protecci\u00f3n del ni\u00f1o y adolescente": "Inversi\u00f3n Protecci\u00f3n",
    "Desarrollo e integraci\u00f3n": "Inversi\u00f3n Desarrollo",
    "Nutrici\u00f3n y alimentaci\u00f3n": "Inversi\u00f3n Nutrici\u00f3n",
}

# All categories for totals
ALL_CATEGORIES = {
    "Educaci\u00f3n": "Inversi\u00f3n Educaci\u00f3n",
    "Salud": "Inversi\u00f3n Salud",
    "Protecci\u00f3n del ni\u00f1o y adolescente": "Inversi\u00f3n Protecci\u00f3n",
    "Desarrollo e integraci\u00f3n": "Inversi\u00f3n Desarrollo",
    "Nutrici\u00f3n y alimentaci\u00f3n": "Inversi\u00f3n Nutrici\u00f3n",
    "Ayuda directa": "Inversi\u00f3n Ayuda Directa",
    "Condiciones de vida": "Inversi\u00f3n Condiciones de Vida",
    "Deportes , recreaci\u00f3n y cultura": "Inversi\u00f3n Deportes y Cultura",
}


def load_excel_data():
    """Load and aggregate investment data from Excel."""
    print(f"Leyendo archivo Excel: {EXCEL_FILE}")
    
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME]
    
    print(f"Hoja: {SHEET_NAME}, Filas: {ws.max_row}")
    
    # Aggregate by year and category using DEVENGADO (column 12)
    # Column indexes: AÑO=1, DEVENGADO=12, Categoria=17
    data = defaultdict(lambda: defaultdict(float))
    
    for i in range(2, ws.max_row + 1):
        year = ws.cell(i, 1).value  # Column A (AÑO)
        categoria = ws.cell(i, 17).value  # Column 17 - Categoria
        devengado = ws.cell(i, 12).value  # Column 12 - DEVENGADO
        
        if year and isinstance(year, (int, float)) and categoria and devengado and isinstance(devengado, (int, float)):
            year = int(year)
            data[year][categoria] += devengado
    
    print(f"Datos agregados para {len(data)} a\u00f1os")
    return data


def generate_sql_inserts(data):
    """Generate SQL INSERT statements for indicadores table."""
    values = []
    
    for year in sorted(data.keys()):
        year_data = data[year]
        
        # Calculate total investment for this year
        total = sum(year_data.values())
        
        # Insert total inversion (in millions of pesos)
        values.append(
            f"('inversion', 'Inversi\u00f3n Total', {total/1e6:.2f}, 'Md', {year}, 'manual')"
        )
        
        # Insert individual categories (only if > 0)
        for categoria_orig, indicador_nombre in ALL_CATEGORIES.items():
            valor = year_data.get(categoria_orig, 0)
            if valor > 0:
                # Values are in pesos, convert to millions (Md)
                values.append(
                    f"('inversion', '{indicador_nombre}', {valor/1e6:.2f}, 'Md', {year}, 'manual')"
                )
    
    if not values:
        print("ERROR: No se generaron valores")
        return ""
    
    sql = f"""
-- Insertar datos de Inversi\u00f3n (manual from Excel)
INSERT INTO indicadores (categoria, indicador_nombre, valor, unidad, periodo, fuente)
VALUES
{','.join(values)}
ON CONFLICT (categoria, indicador_nombre, periodo) DO UPDATE SET
    valor = EXCLUDED.valor,
    fuente = EXCLUDED.fuente;
"""
    return sql


def main():
    """Execute the ETL."""
    print("=" * 60)
    print("ETL Inversi\u00f3n: Cargar datos de Excel a Supabase")
    print("=" * 60)
    
    # Load data
    data = load_excel_data()
    
    if not data:
        print("ERROR: No se pudieron extraer datos del Excel")
        sys.exit(1)
    
    # Show summary
    print("\nResumen de datos extraidos:")
    print("-" * 60)
    years = sorted(data.keys())[-10:]  # Last 10 years
    for year in years:
        total = sum(data[year].values())
        print(f"  {year}: ${total/1e6:.1f}M (total)")
    
    # Generate SQL
    sql = generate_sql_inserts(data)
    
    if not sql:
        print("ERROR: No se genero SQL")
        sys.exit(1)
    
    # Write to file
    output_dir = Path(__file__).parent.resolve() / "output"
    output_dir.mkdir(exist_ok=True)
    output_file = output_dir / "load_inversion_inserts.sql"
    
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(sql)
    
    # Count values
    value_count = sql.count("VALUES") + sql.count("', '") - 2
    
    print("=" * 60)
    print(f"RESUMEN:")
    print(f"  - A\u00f1os procesados: {len(data)}")
    print(f"  - SQL generado: {output_file}")
    print("=" * 60)
    
    # Return SQL for direct execution
    return sql


if __name__ == "__main__":
    sql = main()
    print("\n--- SQL Generado (primeros 2000 caracteres) ---")
    print(sql[:2000])