"""Read an Excel file and return a pandas DataFrame.
If pandas (and its compiled dependencies) are unavailable, fall back to
openpyxl to read the sheet into a list of dicts and then build a minimal
DataFrame‑like object.
"""
import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

def _import_pandas():
    try:
        import pandas as pd
        return pd
    except Exception as e:
        # Pandas failed (often due to missing compiled numpy libs). Log and fall back.
        print(f"[ETL] Pandas import failed: {e}. Falling back to openpyxl.")
        return None

def _fallback_read_excel(path: str):
    """Read specific sheets from Excel file - prioritize BASE DE DATOS."""
    try:
        from openpyxl import load_workbook
        from pandas import DataFrame
        # data_only=True para obtener valores en lugar de formulas
        wb = load_workbook(path, data_only=True)
        
        result = {}
        # Prefer BASE DE DATOS sheet, otherwise first non-empty sheet
        prefer_sheets = ['BASE DE DATOS', 'BASE DE DATOS', '2023', '2022', '2021']
        
        for prefer in prefer_sheets:
            if prefer in wb.sheetnames:
                ws = wb[prefer]
                if ws.max_row > 5:
                    rows = list(ws.iter_rows(values_only=True))
                    # Skip empty rows at start
                    start = 0
                    for i, row in enumerate(rows):
                        if any(row):
                            start = i
                            break
                    if start < len(rows):
                        header = rows[start]
                        data = []
                        for row in rows[start+1:]:
                            if any(row):
                                data.append(dict(zip(header, row)))
                        if data:
                            result[prefer] = DataFrame(data)
                            break
        
        # If no prefered sheet found, load all
        if not result:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row < 5:
                    continue
                rows = list(ws.iter_rows(values_only=True))
                start = 0
                for i, row in enumerate(rows):
                    if any(row):
                        start = i
                        break
                if start < len(rows):
                    header = rows[start]
                    data = []
                    for row in rows[start+1:]:
                        if any(row):
                            data.append(dict(zip(header, row)))
                    if data:
                        result[sheet_name] = DataFrame(data)
        
        return result if result else {"Sheet1": DataFrame()}
    except Exception as fe:
        print(f"[ETL] openpyxl fallback failed: {fe}")
        return {"Sheet1": DataFrame()}

def read_excel(path: str):
    """Read an Excel file and return a dict of sheet names to DataFrames or fallback lists.
    Reads ALL sheets from the workbook and returns them as a dict.
    """
    pd = _import_pandas()
    if pd:
        try:
            # Read all sheets into a dict
            return pd.read_excel(path, sheet_name=None)
        except Exception as e:
            print(f"[ETL] pandas.read_excel error: {e}. Using fallback.")
    # Either pandas not available or read_excel failed
    # Fallback only reads active sheet
    data = _fallback_read_excel(path)
    return {"Sheet1": data}
